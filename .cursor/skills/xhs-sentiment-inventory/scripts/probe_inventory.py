#!/usr/bin/env python3
"""Probe title/author variants and cross-sample judgment conflicts in 帖子全量清单."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path


def norm_title(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).lower()
    text = text.replace("…", "").replace("...", "").replace("·", "")
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", text)


def norm_author(value: object) -> str:
    text = norm_title(value)
    if not text or any(m in text for m in ("裁切", "未知", "未显示", "未识别", "广告")):
        return ""
    return text


def load_inventory(xlsx: Path, sheet: str) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {sheet}. Available: {wb.sheetnames}")
    ws = wb[sheet]
    rows: list[dict] = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not r or r[0] is None:
            continue
        sid = str(int(r[0])) if isinstance(r[0], (int, float)) else str(r[0])
        rows.append(
            {
                "excel_row": i,
                "ID": sid,
                "关键词": str(r[1] or ""),
                "屏号": str(r[2] or ""),
                "位置": str(r[3] or ""),
                "作者": str(r[4] or ""),
                "标题": str(r[5] or ""),
                "内容类型": str(r[6] or ""),
                "判定状态": str(r[7] or ""),
                "是否记入负向": str(r[8] or ""),
                "备注": str(r[9] or "") if len(r) > 9 else "",
            }
        )
    wb.close()
    return rows


def cluster(rows: list[dict]) -> list[dict]:
    by_key: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in rows:
        by_key[(row["关键词"], norm_title(row["标题"]))].append(row)

    groups: list[dict] = []
    for (kw, nt), items in by_key.items():
        by_author: dict[str, list[dict]] = defaultdict(list)
        for row in items:
            by_author[norm_author(row["作者"])].append(row)
        known = [a for a in by_author if a]
        if len(known) <= 1:
            buckets = [items]
        else:
            largest = max(known, key=lambda a: len(by_author[a]))
            by_author[largest].extend(by_author.pop("", []))
            buckets = list(by_author.values())

        for bucket in buckets:
            authors = [x["作者"] for x in bucket]
            titles = [x["标题"] for x in bucket]
            known_authors = [a for a in authors if norm_author(a)]
            judgments = Counter(x["判定状态"] for x in bucket)
            neg = Counter(x["是否记入负向"] for x in bucket)
            title_variants = sorted({t for t in titles if t})
            author_variants = sorted({a for a in known_authors if a})
            unknown_authors = sorted(
                {a for a in authors if a and not norm_author(a)}
            )
            groups.append(
                {
                    "关键词": kw,
                    "norm_title": nt,
                    "rows": len(bucket),
                    "users": len({x["ID"] for x in bucket}),
                    "title_variants": title_variants,
                    "author_variants": author_variants,
                    "unknown_author_labels": unknown_authors,
                    "judgments": dict(judgments),
                    "negativity": dict(neg),
                    "has_text_variant": len(title_variants) > 1
                    or len(author_variants) > 1
                    or (bool(author_variants) and bool(unknown_authors)),
                    "judgment_conflict": len(judgments) > 1,
                    "negativity_conflict": "是" in neg and "否" in neg,
                    "sample_ids": sorted({x["ID"] for x in bucket}, key=lambda s: int(s) if s.isdigit() else s),
                    "excel_rows": [x["excel_row"] for x in bucket],
                }
            )
    return groups


def write_xlsx(groups: list[dict], out: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "摘要"
    variant = wb.create_sheet("标题作者出入")
    conflict = wb.create_sheet("判定冲突")
    neg = wb.create_sheet("负向冲突")

    n = len(groups)
    n_var = sum(1 for g in groups if g["has_text_variant"])
    n_j = sum(1 for g in groups if g["judgment_conflict"])
    n_n = sum(1 for g in groups if g["negativity_conflict"])
    summary.append(["指标", "值"])
    for row in [
        ("同帖组数", n),
        ("标题/作者出入组", n_var),
        ("判定不一致组", n_j),
        ("负向记入冲突组", n_n),
    ]:
        summary.append(list(row))

    headers = [
        "关键词",
        "覆盖用户数",
        "行数",
        "标题变体数",
        "已知作者变体数",
        "标题变体示例",
        "作者变体示例",
        "未知作者标签",
        "判定分布",
        "负向分布",
        "命中ID",
    ]
    for ws, pred in (
        (variant, lambda g: g["has_text_variant"]),
        (conflict, lambda g: g["judgment_conflict"]),
        (neg, lambda g: g["negativity_conflict"]),
    ):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        selected = [g for g in groups if pred(g)]
        selected.sort(key=lambda g: (-g["users"], -g["rows"], g["关键词"]))
        for g in selected:
            ws.append(
                [
                    g["关键词"],
                    g["users"],
                    g["rows"],
                    len(g["title_variants"]),
                    len(g["author_variants"]),
                    " | ".join(g["title_variants"][:6]),
                    " | ".join(g["author_variants"][:6]),
                    " | ".join(g["unknown_author_labels"][:6]),
                    json.dumps(g["judgments"], ensure_ascii=False),
                    json.dumps(g["negativity"], ensure_ascii=False),
                    ",".join(g["sample_ids"][:40]),
                ]
            )

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(
        f"Wrote {out}\n"
        f"groups={n} variants={n_var} judgment_conflicts={n_j} negativity_conflicts={n_n}"
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True, type=Path, help="Workbook with 帖子全量清单")
    parser.add_argument("--sheet", default="帖子全量清单")
    parser.add_argument("--out", required=True, type=Path, help="Output probe xlsx")
    parser.add_argument("--json", type=Path, help="Optional JSON dump of groups")
    args = parser.parse_args()

    rows = load_inventory(args.xlsx, args.sheet)
    groups = cluster(rows)
    write_xlsx(groups, args.out)
    if args.json:
        args.json.parent.mkdir(parents=True, exist_ok=True)
        args.json.write_text(json.dumps(groups, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Wrote {args.json}")


if __name__ == "__main__":
    main()
